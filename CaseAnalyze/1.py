import sys
import os
import re
import bisect
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QFrame, QLabel, QToolBar, QAction, QFileDialog,
    QMessageBox, QSizePolicy, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton
)
from PyQt5.QtCore import Qt, QMimeData
from PyQt5.QtGui import QDrag, QColor, QBrush, QFont, QPalette

import openpyxl
from openpyxl.utils import column_index_from_string

# ------------------- 全局样式表 -------------------
STYLE_SHEET = """
/* 主窗口 */
QMainWindow {
    background-color: #f0f2f5;
}

/* 工具栏 */
QToolBar {
    background-color: #2c3e50;
    border: none;
    padding: 4px;
    spacing: 8px;
}
QToolBar QToolButton {
    background-color: #34495e;
    color: white;
    border: none;
    border-radius: 4px;
    padding: 6px 14px;
    font-weight: bold;
}
QToolBar QToolButton:hover {
    background-color: #3d566e;
}
QToolBar QToolButton:pressed {
    background-color: #1e2b38;
}

/* 标签页 */
QTabWidget::pane {
    border: none;
    background: #f0f2f5;
}
QTabBar::tab {
    background: #e4e7eb;
    padding: 8px 20px;
    margin-right: 2px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    font-weight: bold;
}
QTabBar::tab:selected {
    background: white;
    border-bottom: 3px solid #4A90D9;
}
QTabBar::tab:hover {
    background: #d5d8dd;
}
QTabBar::close-button {
    image: none;  /* 使用自定义关闭符号 */
    subcontrol-position: right;
    subcontrol-origin: padding;
    margin: 0 0 0 8px;
}

/* 总览卡片 */
.OverviewWidget {
    background: transparent;
}
.CaseBlock {
    background: white;
    border-radius: 10px;
    border: 1px solid #d0d7de;
    padding: 10px 15px;
    margin-bottom: 10px;
}
.CaseBlock:hover {
    border-color: #4A90D9;
    box-shadow: 0 2px 8px rgba(74,144,217,0.2);
}
.CaseBlock QLabel#title {
    font-size: 16px;
    font-weight: bold;
    color: #2c3e50;
}
.CaseBlock QLabel#expand {
    font-size: 16px;
    color: #7f8c8d;
}
.CaseBlock .detail-item {
    padding: 2px 0 2px 20px;
    border-left: 3px solid #4A90D9;
    margin-top: 2px;
    font-size: 14px;
}
.CaseBlock .detail-item .dot {
    display: inline-block;
    width: 12px;
    height: 12px;
    border-radius: 3px;
    margin-right: 8px;
}

/* 甘特图表格 */
GanttTableWidget QTableWidget {
    background: white;
    border-radius: 8px;
    border: 1px solid #d0d7de;
    gridline-color: #e9ecef;
    font-size: 13px;
}
GanttTableWidget QTableWidget::item {
    padding: 8px 4px;
}
GanttTableWidget QTableWidget::item:selected {
    background: transparent; /* 取消选中高亮 */
}
GanttTableWidget QHeaderView::section {
    background: #f8f9fa;
    padding: 6px;
    border: 1px solid #dee2e6;
    font-weight: bold;
    font-size: 13px;
}
GanttTableWidget QTableWidget QTableCornerButton::section {
    background: #f8f9fa;
}
"""

# ------------------- 解析器 -------------------
class TestCaseParser:
    # 与之前完全相同，不再重复，保持功能稳定
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.ipn_sheet = self.wb['IPN']
        self.prepost_sheet = self.wb['Pre-Post-Processing']
        self.sub_sheet = self.wb['Substitutions']
        self.trigger_sheet = self.wb['Trigger']

        self.prepost_dict = self._build_dict(self.prepost_sheet, 'A', 'B')
        self.sub_dict = self._build_dict(self.sub_sheet, 'A', 'B')
        self.trigger_dict = self._build_dict(self.trigger_sheet, 'A', 'B')

        header_row = self.ipn_sheet[1]
        self.kategorie_col = self._find_header(header_row, 'Kategorie')
        self.signalname_col = self._find_header(header_row, 'Signalname')
        self.testfallname_col = self._find_header(header_row, 'Testfallname')
        self.octane_col = self._find_header(header_row, 'Octane-ID')
        self.codebeamer_col = self._find_header(header_row, 'Codebeamer-ID')
        self.id_col = self._find_header(header_row, 'ID')
        self.randbedingungen_col = self._find_header(header_row, 'Test_Randbedingungen')

        if self.kategorie_col is None:
            raise ValueError("未找到表头 'Kategorie'")
        if self.signalname_col is None:
            raise ValueError("未找到表头 'Signalname'")
        if self.randbedingungen_col is None:
            raise ValueError("未找到表头 'Test_Randbedingungen'")

    def _find_header(self, header_row, header_name):
        for idx, cell in enumerate(header_row, 1):
            if cell.value and header_name.lower() in str(cell.value).lower():
                return idx
        return None

    def _build_dict(self, sheet, col_key, col_val):
        d = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = row[column_index_from_string(col_key)-1]
            val = row[column_index_from_string(col_val)-1]
            if key is not None and val is not None:
                d[str(key).strip()] = str(val).strip()
        return d

    def parse_all(self):
        cases = []
        for row_idx in range(2, self.ipn_sheet.max_row + 1):
            row = self.ipn_sheet[row_idx]
            kategorie = row[self.kategorie_col-1].value if self.kategorie_col else None
            signalname = row[self.signalname_col-1].value if self.signalname_col else None
            if not kategorie or 'FA' in str(kategorie):
                continue
            if not signalname:
                continue

            case_id = row[self.id_col-1].value if self.id_col else f"Case_{row_idx}"
            testfallname = row[self.testfallname_col-1].value if self.testfallname_col else "未命名"
            octane_id = row[self.octane_col-1].value if self.octane_col else ""
            codebeamer_id = row[self.codebeamer_col-1].value if self.codebeamer_col else ""

            scene_signals = self._parse_kategorie(kategorie)
            action_signals = self._parse_signalname(signalname)
            expect_signals = self._parse_expect_signals(row)

            all_signals = []
            for sig in scene_signals:
                sig['type'] = '场景'
                all_signals.append(sig)
            for sig in action_signals:
                sig['type'] = '动作'
                all_signals.append(sig)
            for sig in expect_signals:
                sig['type'] = '期望'
                all_signals.append(sig)

            for sig in all_signals:
                if 'start' not in sig or sig['start'] is None:
                    sig['start'] = 0
                if 'end' not in sig or sig['end'] is None:
                    sig['end'] = sig['start'] + 1

            case = {
                'id': str(case_id),
                'name': str(testfallname),
                'signals': all_signals,
                'octane_id': str(octane_id),
                'codebeamer_id': str(codebeamer_id)
            }
            cases.append(case)
        return cases

    # 以下解析方法完全相同，省略以节省篇幅（实际使用时需完整复制）
    def _parse_kategorie(self, kategorie_str):
        signals = []
        for part in str(kategorie_str).split(';'):
            part = part.strip()
            if not part:
                continue
            if part in self.prepost_dict:
                content = self.prepost_dict[part]
                for item in self._split_content(content):
                    sig = self._parse_signal_def(item)
                    if sig:
                        signals.append(sig)
            else:
                sig = self._parse_signal_def(part)
                if sig:
                    signals.append(sig)
        return signals

    def _parse_signalname(self, signalname_str):
        signals = []
        text = str(signalname_str).strip()
        if '<<<' in text and '>>>' in text:
            matches = re.findall(r'<<<(.*?)>>>', text)
            for key in matches:
                key = key.strip()
                if key in self.sub_dict:
                    content = self.sub_dict[key]
                    for item in self._split_content(content):
                        sig = self._parse_signal_def(item)
                        if sig:
                            signals.append(sig)
                else:
                    sig = self._parse_signal_def(key)
                    if sig:
                        signals.append(sig)
            trigger_part = re.search(r't=([^,]+(?:,[^,]+)?)', text)
            if trigger_part:
                trigger_expr = trigger_part.group(1)
                times = self._parse_trigger_expr(trigger_expr)
                for sig in signals:
                    if 'start' not in sig and times['start'] is not None:
                        sig['start'] = times['start']
                    if 'end' not in sig and times['end'] is not None:
                        sig['end'] = times['end']
        else:
            for item in self._split_content(text):
                sig = self._parse_signal_def(item)
                if sig:
                    signals.append(sig)
        return signals

    def _parse_expect_signals(self, row):
        signals = []
        for col_idx in range(self.randbedingungen_col, len(row) + 1):
            cell = row[col_idx-1]
            val = cell.value
            if val is None or str(val).strip() in ('rec', 'me', ''):
                continue
            text = str(val).strip()
            header_cell = self.ipn_sheet[1][col_idx-1]
            header_name = header_cell.value if header_cell.value else f"信号{col_idx}"
            if text.startswith('==') or text.startswith('~=='):
                sig = self._parse_signal_def(header_name + text)
                if sig:
                    signals.append(sig)
            else:
                for item in self._split_content(text):
                    sig = self._parse_signal_def(item)
                    if sig:
                        signals.append(sig)
        return signals

    def _split_content(self, content):
        parts = []
        for line in str(content).splitlines():
            for p in line.split('&&'):
                p = p.strip()
                if p:
                    parts.append(p)
        return parts

    def _parse_signal_def(self, def_str):
        match = re.match(r'^([^=><]+?)\s*([=><]=?|>|<)\s*(.+?)(?:,t=(.*))?$', def_str)
        if not match:
            return None
        name = match.group(1).strip()
        op = match.group(2)
        value = match.group(3).strip()
        time_expr = match.group(4) if match.group(4) else None
        sig = {'name': name, 'value': f"{op}{value}"}
        if time_expr:
            times = self._parse_trigger_expr(time_expr)
            sig['start'] = times['start']
            sig['end'] = times['end']
        else:
            sig['start'] = None
            sig['end'] = None
        return sig

    def _parse_trigger_expr(self, expr):
        result = {'start': None, 'end': None}
        expr = expr.strip()
        if ':' in expr:
            start_expr, end_expr = expr.split(':', 1)
            result['start'] = self._resolve_trigger(start_expr)
            result['end'] = self._resolve_trigger(end_expr)
        else:
            result['start'] = self._resolve_trigger(expr)
            result['end'] = None
        return result

    def _resolve_trigger(self, expr):
        expr = expr.strip()
        m = re.match(r'^(\d+(?:\.\d+)?)s$', expr)
        if m:
            return float(m.group(1))
        if expr in self.trigger_dict:
            content = self.trigger_dict[expr]
            return self._resolve_trigger(content)
        if '+' in expr or '-' in expr:
            parts = re.split(r'([+-])', expr)
            if len(parts) >= 3:
                key = parts[0].strip()
                op = parts[1]
                num_str = parts[2].strip()
                base = self._resolve_trigger(key)
                num = float(re.search(r'[\d.]+', num_str).group())
                if op == '+':
                    return base + num
                elif op == '-':
                    return base - num
        return 0.0


# ------------------- PyQt5 界面（美化版） -------------------
class CaseBlock(QFrame):
    def __init__(self, case_data, parent=None):
        super().__init__(parent)
        self.case_data = case_data
        self.is_expanded = False
        self.setObjectName("CaseBlock")
        self.setProperty("class", "CaseBlock")  # 用于样式
        self.setFrameStyle(QFrame.NoFrame)
        self.setAcceptDrops(True)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setMinimumHeight(50)

        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(10, 10, 10, 10)
        self.main_layout.setSpacing(6)

        # 头部
        self.header_widget = QWidget()
        header_layout = QHBoxLayout(self.header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        self.title_label = QLabel(f"{case_data['id']} - {case_data['name']}")
        self.title_label.setObjectName("title")
        self.expand_label = QLabel("▶")
        self.expand_label.setObjectName("expand")
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()
        header_layout.addWidget(self.expand_label)
        self.header_widget.mousePressEvent = self.toggle_expand
        self.main_layout.addWidget(self.header_widget)

        # 详情
        self.detail_widget = QWidget()
        self.detail_widget.setVisible(False)
        detail_layout = QVBoxLayout(self.detail_widget)
        detail_layout.setContentsMargins(0, 4, 0, 0)
        detail_layout.setSpacing(4)
        color_map = {'场景': '#FF9999', '动作': '#99CCFF', '期望': '#99FF99'}
        for sig in case_data['signals']:
            item_widget = QWidget()
            item_layout = QHBoxLayout(item_widget)
            item_layout.setContentsMargins(20, 2, 0, 2)
            dot = QLabel()
            dot.setFixedSize(12, 12)
            dot.setStyleSheet(f"background-color: {color_map.get(sig['type'], '#ccc')}; border-radius: 3px;")
            label = QLabel(f"{sig['type']}: {sig['name']} = {sig.get('value','')} "
                           f"({sig.get('start','?')} → {sig.get('end','?')})")
            label.setStyleSheet("font-size: 13px; color: #2c3e50;")
            item_layout.addWidget(dot)
            item_layout.addWidget(label)
            item_layout.addStretch()
            detail_layout.addWidget(item_widget)
        self.main_layout.addWidget(self.detail_widget)

        self.setCursor(Qt.PointingHandCursor)

    def toggle_expand(self, event):
        self.is_expanded = not self.is_expanded
        self.detail_widget.setVisible(self.is_expanded)
        self.expand_label.setText("▼" if self.is_expanded else "▶")
        self.adjustSize()
        if self.parent():
            self.parent().updateGeometry()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.drag_start_position = event.pos()

    def mouseMoveEvent(self, event):
        if not (event.buttons() & Qt.LeftButton):
            return
        if (event.pos() - self.drag_start_position).manhattanLength() < QApplication.startDragDistance():
            return
        drag = QDrag(self)
        mime = QMimeData()
        mime.setText(self.case_data['id'])
        drag.setMimeData(mime)
        drag.exec_(Qt.MoveAction)

    def dragEnterEvent(self, event):
        event.acceptProposedAction()

    def dropEvent(self, event):
        dragged_id = event.mimeData().text()
        if dragged_id and dragged_id != self.case_data['id']:
            parent_widget = self.parent()
            if hasattr(parent_widget, 'swap_blocks'):
                parent_widget.swap_blocks(dragged_id, self.case_data['id'])
                event.acceptProposedAction()
        event.acceptProposedAction()


class OverviewWidget(QWidget):
    def __init__(self, cases, parent=None):
        super().__init__(parent)
        self.setObjectName("OverviewWidget")
        self.cases = cases
        self.blocks = []
        self.layout = QVBoxLayout(self)
        self.layout.setAlignment(Qt.AlignTop)
        self.layout.setSpacing(10)
        self.layout.setContentsMargins(10, 10, 10, 10)
        self.rebuild_blocks()

    def rebuild_blocks(self):
        for block in self.blocks:
            self.layout.removeWidget(block)
            block.deleteLater()
        self.blocks.clear()
        for case in self.cases:
            block = CaseBlock(case)
            self.blocks.append(block)
            self.layout.addWidget(block)
        self.layout.addStretch()

    def swap_blocks(self, dragged_id, target_id):
        index_drag = None
        index_target = None
        for i, case in enumerate(self.cases):
            if case['id'] == dragged_id:
                index_drag = i
            if case['id'] == target_id:
                index_target = i
        if index_drag is None or index_target is None or index_drag == index_target:
            return
        self.cases[index_drag], self.cases[index_target] = self.cases[index_target], self.cases[index_drag]
        self.rebuild_blocks()

    def expand_all(self, expand):
        for block in self.blocks:
            if block.is_expanded != expand:
                block.toggle_expand(None)


class GanttTableWidget(QWidget):
    def __init__(self, case_data, parent=None):
        super().__init__(parent)
        self.case = case_data
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        title = QLabel(f"{self.case['id']} - {self.case['name']}")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 8px;")
        layout.addWidget(title)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.SolidLine)
        layout.addWidget(self.table)

        self.build_table()

    def build_table(self):
        signals = self.case['signals']
        if not signals:
            self.table.setRowCount(0)
            self.table.setColumnCount(1)
            self.table.setHorizontalHeaderLabels(["无信号"])
            return

        time_points = set()
        for sig in signals:
            if sig.get('start') is not None:
                time_points.add(sig['start'])
            if sig.get('end') is not None:
                time_points.add(sig['end'])
        unique_times = sorted(time_points)
        if not unique_times:
            unique_times = list(range(len(signals)))

        col_labels = [f"{t}s" for t in unique_times]
        self.table.setColumnCount(len(unique_times))
        self.table.setHorizontalHeaderLabels(col_labels)
        self.table.setRowCount(len(signals))
        self.table.setVerticalHeaderLabels([sig['name'] for sig in signals])

        # 设置行高
        for row in range(len(signals)):
            self.table.setRowHeight(row, 35)

        color_map = {'场景': QColor(255, 153, 153),
                     '动作': QColor(153, 204, 255),
                     '期望': QColor(153, 255, 153)}

        for row, sig in enumerate(signals):
            start = sig.get('start')
            end = sig.get('end')
            if start is None:
                start = unique_times[0] if unique_times else 0
            if end is None:
                end = unique_times[-1] if unique_times else start
            col_start = self._find_nearest_index(unique_times, start)
            col_end = self._find_nearest_index(unique_times, end)
            if col_start > col_end:
                col_start, col_end = col_end, col_start
            if col_start <= col_end:
                self.table.setSpan(row, col_start, 1, col_end - col_start + 1)
                item = QTableWidgetItem(sig.get('value', ''))
                item.setTextAlignment(Qt.AlignCenter)
                bg = color_map.get(sig['type'], QColor(200,200,200))
                item.setBackground(QBrush(bg))
                font = item.font()
                font.setBold(True)
                font.setPointSize(11)
                item.setFont(font)
                # 设置前景色深色
                item.setForeground(QBrush(QColor(30, 30, 30)))
                self.table.setItem(row, col_start, item)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()

        # 调整列宽留出边距
        for col in range(self.table.columnCount()):
            self.table.setColumnWidth(col, self.table.columnWidth(col) + 10)

    def _find_nearest_index(self, sorted_list, value):
        if not sorted_list:
            return 0
        try:
            return sorted_list.index(value)
        except ValueError:
            pos = bisect.bisect_left(sorted_list, value)
            if pos == 0:
                return 0
            if pos == len(sorted_list):
                return len(sorted_list)-1
            before = sorted_list[pos-1]
            after = sorted_list[pos]
            if abs(value - before) <= abs(value - after):
                return pos-1
            else:
                return pos


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("测试用例时序分析工具")
        self.setGeometry(100, 100, 1300, 850)

        # 应用样式
        self.setStyleSheet(STYLE_SHEET)

        self.cases_data = []
        self.excel_path = None

        self.create_toolbar()

        self.tab_widget = QTabWidget()
        self.tab_widget.setTabsClosable(True)
        self.tab_widget.tabCloseRequested.connect(self.close_case_tab)
        self.tab_widget.setDocumentMode(True)
        self.setCentralWidget(self.tab_widget)

        self.overview_tab = None
        self.add_overview_tab()

        self.setAcceptDrops(True)
        self.statusBar().showMessage("就绪")

    def create_toolbar(self):
        toolbar = self.addToolBar("工具栏")
        toolbar.setMovable(False)
        # 自定义按钮风格
        btn_style = """
            QToolButton {
                background-color: #34495e;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-weight: bold;
            }
            QToolButton:hover {
                background-color: #3d566e;
            }
            QToolButton:pressed {
                background-color: #1e2b38;
            }
        """
        # 为了使用自定义按钮，改用 QPushButton 添加到工具栏
        # 但工具栏默认使用 QAction，我们可以设置 QAction 的样式，但较复杂，这里用 QToolButton 也支持样式
        # 直接使用 QAction 并应用样式表到 QToolBar
        open_action = QAction("📂 选择 Excel", self)
        open_action.triggered.connect(self.on_open_excel)
        toolbar.addAction(open_action)

        parse_action = QAction("⚙️ 解析", self)
        parse_action.triggered.connect(self.on_parse)
        toolbar.addAction(parse_action)

        export_action = QAction("📊 导出 Excel", self)
        export_action.triggered.connect(self.on_export)
        toolbar.addAction(export_action)

        toolbar.addSeparator()

        expand_all = QAction("📂 全部展开", self)
        expand_all.triggered.connect(lambda: self.batch_expand(True))
        collapse_all = QAction("📁 全部收起", self)
        collapse_all.triggered.connect(lambda: self.batch_expand(False))
        toolbar.addAction(expand_all)
        toolbar.addAction(collapse_all)

        # 应用样式到工具栏的 actions（需要额外设置）
        for action in toolbar.actions():
            widget = toolbar.widgetForAction(action)
            if widget:
                widget.setStyleSheet(btn_style)

    def add_overview_tab(self):
        if self.overview_tab is not None:
            idx = self.tab_widget.indexOf(self.overview_tab)
            if idx != -1:
                self.tab_widget.removeTab(idx)
        self.overview_tab = OverviewWidget(self.cases_data)
        self.tab_widget.insertTab(0, self.overview_tab, "📋 总览")
        self.tab_widget.setCurrentIndex(0)

    def on_open_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.excel_path = file_path
            self.statusBar().showMessage(f"已选择文件: {file_path}")

    def on_parse(self):
        if hasattr(self, 'excel_path') and self.excel_path and os.path.exists(self.excel_path):
            try:
                parser = TestCaseParser(self.excel_path)
                self.cases_data = parser.parse_all()
                self.statusBar().showMessage(f"解析完成，共 {len(self.cases_data)} 个 case")
            except Exception as e:
                QMessageBox.critical(self, "解析错误", f"解析失败: {str(e)}")
                return
        else:
            self.cases_data = self.generate_mock_cases()
            self.statusBar().showMessage("使用模拟数据（未选择 Excel 文件）")

        self.add_overview_tab()
        while self.tab_widget.count() > 1:
            self.tab_widget.removeTab(1)
        for case in self.cases_data:
            self.add_case_gantt_tab(case)
        self.tab_widget.setCurrentIndex(0)

    def generate_mock_cases(self):
        return [
            {
                'id': 'TC-001',
                'name': '启动流程测试',
                'signals': [
                    {'name': '点火信号', 'type': '场景', 'value': 'ON', 'start': 0, 'end': 2},
                    {'name': '油门踏板', 'type': '动作', 'value': '30%', 'start': 1, 'end': 4},
                    {'name': '车速', 'type': '期望', 'value': '50 km/h', 'start': 3, 'end': 5},
                    {'name': '发动机转速', 'type': '期望', 'value': '2000 rpm', 'start': 2, 'end': 6}
                ]
            },
            {
                'id': 'TC-002',
                'name': '制动性能测试',
                'signals': [
                    {'name': '制动主缸压力', 'type': '场景', 'value': '5 bar', 'start': 0, 'end': 3},
                    {'name': '制动踏板行程', 'type': '动作', 'value': '80 mm', 'start': 2, 'end': 5},
                    {'name': '减速度', 'type': '期望', 'value': '-6 m/s²', 'start': 4, 'end': 7},
                    {'name': '轮速', 'type': '期望', 'value': '40 km/h', 'start': 5, 'end': 8}
                ]
            },
            {
                'id': 'TC-003',
                'name': '转向响应测试',
                'signals': [
                    {'name': '方向盘扭矩', 'type': '场景', 'value': '2 Nm', 'start': 0, 'end': 2},
                    {'name': '转向角', 'type': '动作', 'value': '90°', 'start': 1, 'end': 4},
                    {'name': '横摆角速度', 'type': '期望', 'value': '15°/s', 'start': 3, 'end': 6},
                    {'name': '侧向加速度', 'type': '期望', 'value': '4 m/s²', 'start': 4, 'end': 7}
                ]
            }
        ]

    def add_case_gantt_tab(self, case):
        widget = GanttTableWidget(case)
        tab_title = case['id']
        self.tab_widget.addTab(widget, tab_title)

    def close_case_tab(self, index):
        if index == 0:
            return
        self.tab_widget.removeTab(index)

    def batch_expand(self, expand):
        if self.overview_tab:
            self.overview_tab.expand_all(expand)

    def on_export(self):
        QMessageBox.information(self, "导出", "导出功能将生成包含所有 case 甘特图数据的 Excel 文件。")

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path.endswith(('.xlsx', '.xls')):
                self.excel_path = file_path
                self.statusBar().showMessage(f"拖拽文件: {file_path}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # 使用Fusion风格增强跨平台一致性
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
