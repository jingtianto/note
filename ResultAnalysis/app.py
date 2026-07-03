# app.py - 完整版后端
# 运行方式: python app.py
# 依赖: pip install openpyxl flask flask-cors

from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import io
import json
import traceback

app = Flask(__name__)
CORS(app)

# 颜色映射表（fgColor.indexed → 颜色含义）
COLOR_MAP = {
    10: {'name': '红色', 'meaning': 'failed'},
    35: {'name': '浅蓝色', 'meaning': 'signal missing'},
    52: {'name': '橙色', 'meaning': 'trigger not occurred'},
    13: {'name': '黄色', 'meaning': 'warning'},
    11: {'name': '绿色', 'meaning': 'passed'},
    9:  {'name': 'blank', 'meaning': 'blank'}
}


def safe_int(value):
    """安全地将 openpyxl 对象转为 Python int"""
    if value is None:
        return None
    if hasattr(value, 'value'):
        val = value.value
        if val is None:
            return None
        try:
            return int(val)
        except (TypeError, ValueError):
            return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def get_cell_color(cell):
    """获取单元格颜色索引"""
    color_indexed = None
    fill = cell.fill
    if fill is not None:
        if hasattr(fill, 'fgColor') and fill.fgColor is not None:
            fg = fill.fgColor
            if hasattr(fg, 'indexed'):
                idx = fg.indexed
                if idx is not None and idx != -1:
                    color_indexed = safe_int(idx)
        if color_indexed is None and hasattr(fill, 'bgColor') and fill.bgColor is not None:
            bg = fill.bgColor
            if hasattr(bg, 'indexed'):
                idx = bg.indexed
                if idx is not None and idx != -1:
                    color_indexed = safe_int(idx)
    return color_indexed


def get_cell_value(cell):
    """获取单元格值"""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, (int, float, str, bool)):
        return val
    if hasattr(val, 'value'):
        return val.value
    return str(val)


def find_keyword_column(sheet, keyword_text, search_range=None):
    """在指定范围内查找关键字所在的列和行"""
    min_row = 1
    max_row = sheet.max_row

    if search_range:
        if search_range.get('type') == 'rows':
            min_row = search_range.get('start_row', 1)
            max_row = search_range.get('end_row', sheet.max_row)
        elif search_range.get('type') == 'row':
            row_num = search_range.get('row_number', 1)
            min_row = row_num
            max_row = row_num

    for row_idx in range(min_row, max_row + 1):
        row = sheet[row_idx]
        for col_idx, cell in enumerate(row, 1):
            if cell.value and keyword_text in str(cell.value):
                return col_idx, row_idx
    return None, None


@app.route('/api/analyze', methods=['POST'])
def analyze():
    """接收文件列表和方案，返回所有文件的统计结果"""
    try:
        files = request.files.getlist('files')
        if not files:
            return jsonify({'success': False, 'error': '未上传文件'}), 400

        scheme_json = request.form.get('scheme')
        if not scheme_json:
            return jsonify({'success': False, 'error': '未提供方案'}), 400

        scheme = json.loads(scheme_json)

        results = []
        for file in files:
            try:
                file_bytes = file.read()
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                result = process_workbook(wb, scheme)
                results.append({
                    'filename': file.filename,
                    'success': result.get('success', False),
                    'data': result.get('data', {}),
                    'error': result.get('error', None)
                })
            except Exception as e:
                results.append({
                    'filename': file.filename,
                    'success': False,
                    'error': str(e)
                })

        return jsonify({'success': True, 'results': results})

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'trace': traceback.format_exc()
        }), 500


def process_workbook(workbook, scheme):
    """处理单个工作簿"""
    keywords = scheme.get('keywords', [])
    base_key_idx = scheme.get('baseKey')
    if base_key_idx is not None and base_key_idx != '':
        base_key_idx = int(base_key_idx)
    else:
        base_key_idx = -1

    # 获取每个关键字的列和行
    kw_infos = []
    for idx, kw in enumerate(keywords):
        sheet_name = kw.get('sheet', '')
        if sheet_name not in workbook.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" 不存在'}

        sheet = workbook[sheet_name]
        search_range = None
        if kw.get('searchRange') == 'rows':
            search_range = {'type': 'rows', 'start_row': kw.get('startRow'), 'end_row': kw.get('endRow')}
        elif kw.get('searchRange') == 'row':
            search_range = {'type': 'row', 'row_number': kw.get('rowNumber')}

        col, row_num = find_keyword_column(sheet, kw.get('keywordText', ''), search_range)
        if col is None:
            return {'success': False, 'error': f'在Sheet "{sheet_name}" 中未找到关键字 "{kw.get("keywordText")}"'}

        target_col = col
        if kw.get('valuePosition') == 'left':
            target_col = col - (kw.get('offset', 1))
        elif kw.get('valuePosition') == 'right':
            target_col = col + (kw.get('offset', 1))
        if target_col < 1:
            target_col = 1

        kw_infos.append({
            'sheet': sheet,
            'sheet_name': sheet_name,
            'col': target_col,
            'row_num': row_num,
            'kw': kw,
            'index': idx
        })

    # ---- cell 模式 ----
    cell_data = []
    for info in kw_infos:
        if info['kw'].get('collectType') != 'cell':
            continue
        cell = info['sheet'].cell(info['row_num'], info['col'])
        val = get_cell_value(cell)
        color_indexed = None
        if info['kw'].get('collectColor'):
            color_indexed = get_cell_color(cell)

        valid = True
        if info['kw'].get('excludeEmpty') and (val is None or val == ''):
            valid = False
        if valid and info['kw'].get('excludeValues'):
            exclude_vals = info['kw'].get('excludeValues', [])
            if val is not None and str(val).strip() in [str(e).strip() for e in exclude_vals]:
                valid = False
        if valid:
            color_info = COLOR_MAP.get(color_indexed, None)
            cell_data.append({
                'keyword': info['kw'],
                'value': val,
                'color_indexed': color_indexed,
                'color_name': color_info['name'] if color_info else None,
                'color_meaning': color_info['meaning'] if color_info else None
            })

    # ---- column 模式 ----
    col_infos = [info for info in kw_infos if info['kw'].get('collectType') == 'column']
    column_data = {'rows': [], 'keywords': [info['kw'] for info in col_infos]}

    if col_infos:
        all_row_numbers = set()
        for info in col_infos:
            sheet = info['sheet']
            for row_idx in range(1, sheet.max_row + 1):
                row = sheet[row_idx]
                if row.hidden:
                    continue
                if row_idx == info['row_num']:
                    continue
                all_row_numbers.add(row_idx)

        for row_idx in all_row_numbers:
            row_values = []
            valid = True
            for info in col_infos:
                cell = info['sheet'].cell(row_idx, info['col'])
                val = get_cell_value(cell)
                color_indexed = None
                if info['kw'].get('collectColor'):
                    color_indexed = get_cell_color(cell)

                if base_key_idx >= 0 and info['index'] == base_key_idx:
                    if val is None or val == '':
                        valid = False
                        break

                if info['kw'].get('excludeEmpty') and (val is None or val == ''):
                    valid = False
                    break
                if info['kw'].get('excludeValues'):
                    exclude_vals = info['kw'].get('excludeValues', [])
                    if val is not None and str(val).strip() in [str(e).strip() for e in exclude_vals]:
                        valid = False
                        break

                color_info = COLOR_MAP.get(color_indexed, None)
                row_values.append({
                    'value': val,
                    'color_indexed': color_indexed,
                    'color_name': color_info['name'] if color_info else None,
                    'color_meaning': color_info['meaning'] if color_info else None
                })

            if valid:
                column_data['rows'].append(row_values)

    return {
        'success': True,
        'data': {
            'cell_data': cell_data,
            'column_data': column_data,
            'has_cell': len(cell_data) > 0,
            'has_column': len(column_data['rows']) > 0,
            'keyword_labels': [kw.get('keywordText', '') for kw in keywords]
        }
    }


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    print('🚀 启动 Excel 分析服务...')
    print('📡 http://127.0.0.1:5000')
    print('按 Ctrl+C 停止服务')
    app.run(debug=False, host='127.0.0.1', port=5000)