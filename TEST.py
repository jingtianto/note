import openpyxl
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import re
import io
from openpyxl.styles import PatternFill

# ---------------------- 配置区 ----------------------
INPUT_FILE = r"D:\PythonProject\自动分析测试结果\1.xlsx"
OUTPUT_FILE = r"D:\PythonProject\自动分析测试结果\2.xlsx"

# 精准匹配批注格式 + 保留小数精度
CELL_VALUE_PATTERN = re.compile(r'(~?)==?\s*(-?\d+\.?\d*)[dD]', re.DOTALL)
NOTE_MIN_PATTERN = re.compile(r'Minimum value:\s*(-?\d+\.?\d*)', re.DOTALL)
NOTE_MAX_PATTERN = re.compile(r'Maximum value:\s*(-?\d+\.?\d*)', re.DOTALL)
NOTE_AVG_PATTERN = re.compile(r'Average:\s*(-?\d+\.?\d*)', re.DOTALL)
TIME_FAILED_PATTERN = re.compile(r't\d*=\[(\d+\.\d+),(\d+\.\d+)\]\s*failed', re.DOTALL)

def rgb_to_hex(rgb):
    """将openpyxl的RGB元组(0-255)转换为matplotlib可用的十六进制颜色"""
    if not rgb:
        return '#000000'  # 默认黑色
    # 处理RGBA（如果有透明度，忽略）
    if len(rgb) == 4:
        rgb = rgb[:3]
    return '#{:02x}{:02x}{:02x}'.format(*rgb)

def get_cell_fill_color(cell):
    """获取单元格填充颜色（返回十六进制），无填充返回黑色"""
    fill = cell.fill
    # 🔥 修复：兼容openpyxl不同版本的填充格式
    if fill.patternType is None or fill.patternType == 'none' or fill.fgColor.rgb is None:
        return '#000000'  # 无填充→黑色
    
    # 处理不同的RGB格式
    rgb = fill.fgColor.rgb
    if isinstance(rgb, str):
        # 处理ARGB字符串（Excel默认格式，前两位是透明度）
        if len(rgb) == 8:
            rgb_hex = rgb[2:]  # 去掉透明度，保留RGB
            return f'#{rgb_hex}'
        elif len(rgb) == 6:
            return f'#{rgb}'
    elif isinstance(rgb, tuple):
        # 处理RGB元组
        return rgb_to_hex(rgb)
    
    # 兜底：如果解析失败，打印调试信息并返回黑色
    print(f"⚠️  颜色解析失败: {rgb}，使用黑色")
    return '#000000'

def parse_cell_value(cell_text):
    """解析期望值 + 波浪线逻辑"""
    if not cell_text:
        return None, False
    cell_text_str = str(cell_text).replace('\n','').replace('\r','').strip()
    if cell_text_str in ('me', 'None', ''):
        return None, False

    match = CELL_VALUE_PATTERN.search(cell_text_str)
    if match:
        is_dashed = bool(match.group(1))
        value = float(match.group(2))
        return value, is_dashed
    return None, False

def parse_note_data(note_text):
    """精准解析批注 + 保留小数精度"""
    if not note_text:
        return 0.0, 0.0, 0.0, []

    note_text = note_text.replace('\n',' ').replace('\r',' ').strip()
    print(f"📝 解析批注: {note_text}")

    # 提取数值
    min_match = NOTE_MIN_PATTERN.search(note_text)
    min_val = float(min_match.group(1)) if min_match else 0.0
    max_match = NOTE_MAX_PATTERN.search(note_text)
    max_val = float(max_match.group(1)) if max_match else 0.0
    avg_match = NOTE_AVG_PATTERN.search(note_text)
    avg_val = float(avg_match.group(1)) if avg_match else 0.0

    # 提取时间并保留三位小数
    time_matches = TIME_FAILED_PATTERN.findall(note_text)
    failed_intervals = []
    for t_start, t_end in time_matches:
        t_start = round(float(t_start), 3)
        t_end = round(float(t_end), 3)
        failed_intervals.append((t_start, t_end))
    
    print(f"   → Min:{min_val}, Max:{max_val}, Avg:{avg_val}, 时间区间:{failed_intervals}")
    return min_val, max_val, avg_val, failed_intervals

def process_test_case(ws_in, row_idx, wb_out):
    """处理单个测试用例（单个ID），创建独立Sheet"""
    # 获取ID（去重+格式化）
    test_id = ws_in.cell(row_idx, 1).value 
    if not test_id:
        print(f"⚠️  第{row_idx}行无ID，跳过")
        return
    
    test_id_str = str(test_id).strip()
    sheet_name = f"TC_{test_id_str}"
    
    # 如果Sheet已存在，先删除（避免重复）
    if sheet_name in wb_out.sheetnames:
        del wb_out[sheet_name]
    ws_out = wb_out.create_sheet(title=sheet_name)
    
    print(f"\n========== 处理测试用例 {test_id_str}（第{row_idx}行）==========")
    valid_signals = []

    # 遍历所有列（3列到最后一列）
    for col_idx in range(3, ws_in.max_column + 1):
        # 获取信号名（表头行）
        signal_name_cell_header = ws_in.cell(1, col_idx)
        signal_name = signal_name_cell_header.value or f"Signal_Col{col_idx}"
        
        # 获取数据单元格（当前行对应列，🔥 核心修复：读取数据行的填充色）
        data_cell = ws_in.cell(row_idx, col_idx)
        # 🔥 关键修改：读取数据行单元格的填充色（而非表头行）
        signal_color = get_cell_fill_color(data_cell)
        print(f"🎨 信号 {signal_name[:30]} (列{col_idx}, 行{row_idx}) 颜色: {signal_color}")
        
        cell_val = data_cell.value
        note_text = data_cell.comment.text if (data_cell.comment and data_cell.comment.text) else ""

        if not cell_val or str(cell_val).strip() in ('me', 'None', ''):
            continue

        target_value, is_dashed = parse_cell_value(cell_val)
        min_val, max_val, avg_val, failed_intervals = parse_note_data(note_text)

        if target_value is None and not failed_intervals:
            continue

        if target_value is None:
            target_value = -2.0 if "-2d" in str(cell_val) else 0.0

        valid_signals.append({
            "name": signal_name,
            "color": signal_color,  # 数据行的填充色
            "target": target_value,
            "is_dashed": is_dashed,
            "min": min_val,
            "max": max_val,
            "avg": avg_val,
            "times": failed_intervals
        })
        print(f"✅ 有效信号: {signal_name[:50]} (列{col_idx})")

    if not valid_signals:
        print(f"⚠️  测试用例{test_id_str}无有效信号")
        return

    # 动态计算X轴范围
    all_times = []
    for sig in valid_signals:
        for t_start, t_end in sig["times"]:
            all_times.append(t_start)
            all_times.append(t_end)
    x_min = min(all_times) - 0.05 if all_times else 0
    x_max = max(all_times) + 0.05 if all_times else 10

    # 绘制图表
    n = len(valid_signals)
    fig, axes = plt.subplots(n, 1, figsize=(20, 4*n), sharex=True)
    if n == 1:
        axes = [axes]

    for i, sig in enumerate(valid_signals):
        ax = axes[i]
        
        # 🔥 应用数据行的填充色到信号名
        ax.set_title(sig["name"], 
                     fontweight='bold', 
                     fontsize=9,  # 加大字号，颜色更醒目
                     loc='left', 
                     pad=10,
                     color=sig["color"])  # 关键：使用数据行的颜色
        ax.grid(True, alpha=0.3)
        ax.set_ylabel("Value", fontsize=10)

        # Y轴自适应（包含所有值：min/max/avg/target）
        all_y_values = [sig["min"], sig["max"], sig["avg"], sig["target"]]
        y_min = min(all_y_values) - 0.5
        y_max = max(all_y_values) + 0.5
        ax.set_ylim(y_min, y_max)
        print(f"📌 信号 {sig['name'][:20]} Y轴范围: {y_min} ~ {y_max} (Avg:{sig['avg']})")

        # 第三层：Actual（实际值区间）
        actual_handles = []
        actual_labels = []
        for t_start, t_end in sig["times"]:
            line_min = ax.hlines(sig["min"], t_start, t_end, color='gold', linewidth=3, 
                                label=f"Actual Min Value", zorder=1)
            line_max = ax.hlines(sig["max"], t_start, t_end, color='gold', linewidth=3, 
                                label=f"Actual Max Value", zorder=1)
            fill_range = ax.fill_between([t_start, t_end], sig["min"], sig["max"], 
                                        color='gold', alpha=0.5, label=f"Actual Range", zorder=1)
            actual_handles.extend([line_min, line_max, fill_range])
            actual_labels.extend([f"Actual Min Value", f"Actual Max Value", f"Actual Range"])

        # 第二层：Expected（期望值）
        expected_handles = []
        expected_labels = []
        for t_start, t_end in sig["times"]:
            line_exp = ax.hlines(sig["target"], t_start, t_end, color='blue', 
                                linestyle='--' if sig["is_dashed"] else '-', 
                                linewidth=2, label=f"Expected: {sig['target']}", zorder=2)
            expected_handles.append(line_exp)
            expected_labels.append(f"Expected: {sig['target']}")

        # 第一层：Average（平均值）
        avg_handles = []
        avg_labels = []
        for t_start, t_end in sig["times"]:
            line_avg = ax.plot((t_start + t_end)/2, sig["avg"], 'yo', markersize=12, 
                             label=f"Actual Average: {sig['avg']}", zorder=3)[0]
            avg_handles.append(line_avg)
            avg_labels.append(f"Actual Average: {sig['avg']}")
            print(f"📍 Average坐标: x={(t_start + t_end)/2}, y={sig['avg']}")

        # 子图右侧独立图例
        all_handles = actual_handles + expected_handles + avg_handles
        all_labels = actual_labels + expected_labels + avg_labels
        by_label = dict(zip(all_labels, all_handles))  # 去重
        
        ax.legend(by_label.values(), by_label.keys(),
                  loc='upper left',
                  bbox_to_anchor=(1.01, 1.0),
                  fontsize=8,
                  frameon=True,
                  borderaxespad=0)

        # 调整子图内边距
        ax.set_position([ax.get_position().x0, ax.get_position().y0, 
                        ax.get_position().width*0.85, ax.get_position().height])

    # X轴优化
    axes[-1].set_xlabel("Time (s)", fontsize=12)
    axes[-1].set_xlim(x_min, x_max)
    axes[-1].xaxis.set_major_formatter(plt.FormatStrFormatter('%.3f'))

    # 整体布局调整
    plt.subplots_adjust(top=0.95, bottom=0.05, hspace=0.5)
    fig.suptitle(f"Test Case {test_id_str} - Signal Analysis", fontsize=16, y=0.98)

    # 保存图表到Sheet
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', dpi=300)
    buf.seek(0)
    ws_out.add_image(Image(buf), "A1")
    plt.close()

# ---------------------- 主程序：批量处理所有ID ----------------------
if __name__ == "__main__":
    # 加载输入文件（🔥 修复：保留单元格格式）
    wb_in = openpyxl.load_workbook(INPUT_FILE, data_only=False)  # 关键：data_only=False 保留格式
    ws_in = wb_in.active
    wb_out = openpyxl.Workbook()

    # 遍历A列所有非空行（从第2行开始，跳过表头）
    processed_ids = set()  # 记录已处理的ID，避免重复
    for row_idx in range(2, ws_in.max_row + 1):
        current_id = ws_in.cell(row_idx, 1).value
        if not current_id:
            continue
        
        current_id_str = str(current_id).strip()
        # 跳过重复ID（同一ID只处理一次）
        if current_id_str in processed_ids:
            continue
        
        # 处理当前ID，创建独立Sheet
        process_test_case(ws_in, row_idx, wb_out)
        processed_ids.add(current_id_str)

    # 清理默认Sheet
    if "Sheet" in wb_out.sheetnames:
        del wb_out["Sheet"]

    # 保存输出文件
    wb_out.save(OUTPUT_FILE)
    print(f"\n🎉 全部处理完成！文件已保存到: {OUTPUT_FILE}")
    print(f"📋 处理结果：共识别并处理 {len(processed_ids)} 个唯一测试用例ID")
    print(f"📑 生成的Sheet页：{list(wb_out.sheetnames)}")