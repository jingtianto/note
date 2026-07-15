import sys
import re
import os
import traceback
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QFileDialog, QTabWidget, QLineEdit, QLabel,
    QListWidget, QAbstractItemView, QMessageBox, QTextEdit,
    QGroupBox, QFrame
)
from PyQt5.QtCore import Qt, QUrl, QTimer
from PyQt5.QtGui import QColor, QBrush, QDragEnterEvent, QDropEvent, QFont
import openpyxl
from openpyxl.styles import PatternFill

# ---------- 增强版解析函数 ----------
def parse_dtc_text(text):
    """
    解析与dtc.txt格式相同的文本，兼容多种变体。
    返回：{'main': [(id, short, desc, pre, before, post), ...], 'perf': [...]}
    """
    lines = text.splitlines()
    phases = ['Preprocess', 'Before Stimulation', 'Postprocess']
    phase_data = {phase: {'main': [], 'perf': []} for phase in phases}
    
    current_phase = None
    current_module = None
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        # 检测阶段标题 (如 "(514.81480) Preprocess")
        m = re.match(r'^\(\d+\.\d+\)\s+(.+)$', line)
        if m:
            current_phase = m.group(1).strip()
            i += 1
            continue
        
        # 检测模块 (忽略大小写)
        if line.upper().startswith('IPN_MAIN'):
            current_module = 'main'
            i += 1
            continue
        elif line.upper().startswith('IPN_PERF'):
            current_module = 'perf'
            i += 1
            continue
        
        # 检测DTC条目 (0x开头的6位十六进制)
        if re.match(r'^0x[0-9A-Fa-f]{6}$', line):
            dtc_id = line
            short_name = ''
            desc = ''
            status = ''
            j = i + 1
            # 在DTC块中读取键值对
            while j < len(lines):
                line_j = lines[j].strip()
                if not line_j:
                    j += 1
                    continue
                # 遇到下一个DTC或新模块/阶段则停止
                if re.match(r'^0x[0-9A-Fa-f]{6}$', line_j):
                    break
                if line_j.upper().startswith('IPN_MAIN') or line_j.upper().startswith('IPN_PERF') or re.match(r'^\(\d+\.\d+\)', line_j):
                    break
                
                # 处理键：ShortName, Description, Status (支持同一行或下一行)
                def extract_value(line, key):
                    if line.startswith(key):
                        # 尝试分割
                        parts = line.split(':', 1) if ':' in line else line.split(' ', 1)
                        if len(parts) > 1:
                            return parts[1].strip()
                        else:
                            return None
                    return None
                
                val = None
                if line_j.startswith('ShortName'):
                    val = extract_value(line_j, 'ShortName')
                    if val is None and j + 1 < len(lines):
                        val = lines[j+1].strip()
                        j += 1
                    if val:
                        short_name = val
                    j += 1
                    continue
                elif line_j.startswith('Description'):
                    val = extract_value(line_j, 'Description')
                    if val is None and j + 1 < len(lines):
                        val = lines[j+1].strip()
                        j += 1
                    if val:
                        desc = val
                    j += 1
                    continue
                elif line_j.startswith('Status'):
                    val = extract_value(line_j, 'Status')
                    if val is None and j + 1 < len(lines):
                        val = lines[j+1].strip()
                        j += 1
                    if val:
                        status = val
                    j += 1
                    continue
                else:
                    j += 1
                    continue
            
            if current_phase and current_module and current_phase in phase_data:
                phase_data[current_phase][current_module].append((dtc_id, short_name, desc, status))
            i = j
            continue
        i += 1

    # 构建DTC顺序（按Preprocess->Before->Postprocess）
    order_main = []
    order_perf = []
    for phase in phases:
        for module in ['main', 'perf']:
            for dtc_id, _, _, _ in phase_data[phase][module]:
                if module == 'main' and dtc_id not in order_main:
                    order_main.append(dtc_id)
                elif module == 'perf' and dtc_id not in order_perf:
                    order_perf.append(dtc_id)
    
    # 整合三个阶段的状态
    file_data = {'main': {}, 'perf': {}}
    for phase in phases:
        for module in ['main', 'perf']:
            for dtc_id, short, desc, status in phase_data[phase][module]:
                if dtc_id not in file_data[module]:
                    file_data[module][dtc_id] = {'short': short, 'desc': desc, 'pre': '', 'before': '', 'post': ''}
                if phase == 'Preprocess':
                    file_data[module][dtc_id]['pre'] = status
                elif phase == 'Before Stimulation':
                    file_data[module][dtc_id]['before'] = status
                elif phase == 'Postprocess':
                    file_data[module][dtc_id]['post'] = status
    
    main_list = []
    for dtc in order_main:
        info = file_data['main'][dtc]
        main_list.append((dtc, info['short'], info['desc'], info['pre'], info['before'], info['post']))
    
    perf_list = []
    for dtc in order_perf:
        info = file_data['perf'][dtc]
        perf_list.append((dtc, info['short'], info['desc'], info['pre'], info['before'], info['post']))
    
    return {'main': main_list, 'perf': perf_list}


class FileItem:
    def __init__(self, name, data):
        self.name = name
        self.data = data


class FrozenTableWidget(QWidget):
    """实现冻结列的表格组件，左侧固定列数可配置"""
    def __init__(self, freeze_cols=3, parent=None):
        super().__init__(parent)
        self.freeze_cols = freeze_cols
        self.initUI()
        self.setStyleSheet("QTableWidget { border: none; }")
        
    def initUI(self):
        layout = QHBoxLayout(self)
        layout.setSpacing(0)
        layout.setContentsMargins(0, 0, 0, 0)
        
        self.left_table = QTableWidget()
        self.left_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_table.setSortingEnabled(False)
        self.left_table.setFocusPolicy(Qt.NoFocus)
        
        self.right_table = QTableWidget()
        self.right_table.setSortingEnabled(False)
        
        # 同步滚动
        self.right_table.verticalScrollBar().valueChanged.connect(
            lambda val: self.left_table.verticalScrollBar().setValue(val)
        )
        self.left_table.verticalScrollBar().valueChanged.connect(
            lambda val: self.right_table.verticalScrollBar().setValue(val)
        )
        
        # 同步选择
        self.left_table.itemSelectionChanged.connect(self.sync_selection)
        self.right_table.itemSelectionChanged.connect(self.sync_selection)
        
        layout.addWidget(self.left_table)
        layout.addWidget(self.right_table)
        
    def sync_selection(self):
        selected_rows = set()
        for table in [self.left_table, self.right_table]:
            for item in table.selectedItems():
                selected_rows.add(item.row())
        if not selected_rows:
            return
        row = next(iter(selected_rows))
        for table in [self.left_table, self.right_table]:
            table.selectRow(row)
            
    def setRowCount(self, rows):
        self.left_table.setRowCount(rows)
        self.right_table.setRowCount(rows)
        
    def setColumnCount(self, cols):
        left_cols = min(self.freeze_cols, cols)
        right_cols = max(0, cols - self.freeze_cols)
        self.left_table.setColumnCount(left_cols)
        self.right_table.setColumnCount(right_cols)
        
    def setHorizontalHeaderLabels(self, labels):
        # 由于使用双行表头，此处不设置标准表头，而是通过数据区域实现
        pass
        
    def setItem(self, row, col, item):
        if col < self.freeze_cols:
            self.left_table.setItem(row, col, item)
        else:
            self.right_table.setItem(row, col - self.freeze_cols, item)
            
    def item(self, row, col):
        if col < self.freeze_cols:
            return self.left_table.item(row, col)
        else:
            return self.right_table.item(row, col - self.freeze_cols)
            
    def clear(self):
        self.left_table.clear()
        self.right_table.clear()
        
    def resizeColumnsToContents(self):
        self.left_table.resizeColumnsToContents()
        self.right_table.resizeColumnsToContents()
        
    def setColumnWidth(self, col, width):
        if col < self.freeze_cols:
            self.left_table.setColumnWidth(col, width)
        else:
            self.right_table.setColumnWidth(col - self.freeze_cols, width)
            
    def columnWidth(self, col):
        if col < self.freeze_cols:
            return self.left_table.columnWidth(col)
        else:
            return self.right_table.columnWidth(col - self.freeze_cols)
            
    def rowCount(self):
        return self.left_table.rowCount()
        
    def columnCount(self):
        return self.left_table.columnCount() + self.right_table.columnCount()
        
    def setRowHidden(self, row, hidden):
        self.left_table.setRowHidden(row, hidden)
        self.right_table.setRowHidden(row, hidden)
        
    def isRowHidden(self, row):
        return self.left_table.isRowHidden(row)
        
    def setSpan(self, row, col, rowSpan, colSpan):
        if col >= self.freeze_cols:
            self.right_table.setSpan(row, col - self.freeze_cols, rowSpan, colSpan)
        else:
            self.left_table.setSpan(row, col, rowSpan, colSpan)
    
    def setRowHeight(self, row, height):
        self.left_table.setRowHeight(row, height)
        self.right_table.setRowHeight(row, height)
            
    def setFont(self, font):
        self.left_table.setFont(font)
        self.right_table.setFont(font)
        
    def setSortingEnabled(self, enabled):
        self.left_table.setSortingEnabled(enabled)
        self.right_table.setSortingEnabled(enabled)
        
    def setAlternatingRowColors(self, enabled):
        self.left_table.setAlternatingRowColors(enabled)
        self.right_table.setAlternatingRowColors(enabled)
        
    def verticalHeader(self):
        return self.left_table.verticalHeader()
        
    def horizontalHeader(self):
        return self.left_table.horizontalHeader()


class DtcCompareApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DTC 对比分析工具")
        self.setGeometry(100, 100, 1200, 700)
        self.files = []
        self.initUI()
        self.apply_styles()

    def initUI(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(6)
        main_layout.setContentsMargins(8, 8, 8, 8)

        # ----- 顶部工具栏 -----
        tool_frame = QFrame()
        tool_frame.setFrameShape(QFrame.StyledPanel)
        tool_layout = QHBoxLayout(tool_frame)
        tool_layout.setSpacing(8)

        self.btn_add = QPushButton("📂 加载文件")
        self.btn_add.clicked.connect(self.add_files)
        self.btn_add.setToolTip("添加 .txt 或 .xlsx 文件")
        self.btn_remove = QPushButton("🗑 移除选中")
        self.btn_remove.clicked.connect(self.remove_selected)
        self.btn_remove.setToolTip("移除列表中选中的文件")
        self.btn_clear = QPushButton("🧹 清空全部")
        self.btn_clear.clicked.connect(self.clear_all)
        self.btn_clear.setToolTip("清空所有文件和粘贴内容")

        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.file_list.setMaximumHeight(80)
        self.file_list.setMinimumHeight(60)

        tool_layout.addWidget(self.btn_add)
        tool_layout.addWidget(self.btn_remove)
        tool_layout.addWidget(self.btn_clear)
        tool_layout.addWidget(self.file_list, 1)

        main_layout.addWidget(tool_frame)

        # ----- 粘贴区域 (折叠) -----
        paste_group = QGroupBox("▼ 粘贴解析")
        paste_group.setCheckable(True)
        paste_group.setChecked(False)
        paste_group.setMaximumHeight(30 if not paste_group.isChecked() else 120)
        paste_group.toggled.connect(lambda checked: paste_group.setMaximumHeight(120 if checked else 30))
        paste_layout = QVBoxLayout(paste_group)
        self.paste_edit = QTextEdit()
        self.paste_edit.setPlaceholderText("在此粘贴 dtc.txt 格式内容... 多个文件用 '---' 分隔")
        self.paste_edit.setMaximumHeight(60)
        self.btn_paste_parse = QPushButton("▶ 解析粘贴")
        self.btn_paste_parse.clicked.connect(self.parse_paste)
        paste_layout.addWidget(self.paste_edit)
        paste_layout.addWidget(self.btn_paste_parse)
        main_layout.addWidget(paste_group)

        # ----- 搜索栏 -----
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 搜索:"))
        self.search_line = QLineEdit()
        self.search_line.setPlaceholderText("输入 DTC 编号或描述关键字...")
        self.search_line.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_line)
        main_layout.addLayout(search_layout)

        # ----- 标签页 (总览 + 单文件) -----
        self.tab_widget = QTabWidget()
        self.tab_widget.setTabsClosable(False)
        self.overview_tab = QWidget()
        self.setup_overview_tab()
        self.tab_widget.addTab(self.overview_tab, "📊 总览对比")
        main_layout.addWidget(self.tab_widget, 1)

        # ----- 底部操作栏 -----
        bottom_layout = QHBoxLayout()
        self.btn_export = QPushButton("💾 导出当前页面为 Excel")
        self.btn_export.clicked.connect(self.export_current)
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.btn_export)
        main_layout.addLayout(bottom_layout)

        self.statusBar().showMessage("就绪")
        self.setAcceptDrops(True)

    def setup_overview_tab(self):
        layout = QVBoxLayout(self.overview_tab)
        layout.setSpacing(4)
        # 使用冻结表格，固定前3列
        self.overview_main_table = FrozenTableWidget(freeze_cols=3)
        self.overview_perf_table = FrozenTableWidget(freeze_cols=3)
        font = QFont("Segoe UI", 9)
        self.overview_main_table.setFont(font)
        self.overview_perf_table.setFont(font)
        self.overview_main_table.setAlternatingRowColors(True)
        self.overview_perf_table.setAlternatingRowColors(True)
        # 隐藏垂直表头
        self.overview_main_table.left_table.verticalHeader().setVisible(False)
        self.overview_main_table.right_table.verticalHeader().setVisible(False)
        self.overview_perf_table.left_table.verticalHeader().setVisible(False)
        self.overview_perf_table.right_table.verticalHeader().setVisible(False)
        layout.addWidget(QLabel("IPN_MAIN"))
        layout.addWidget(self.overview_main_table)
        layout.addWidget(QLabel("IPN_PERF"))
        layout.addWidget(self.overview_perf_table)

    def apply_styles(self):
        style = """
            QMainWindow { background-color: #f5f5f5; }
            QFrame {
                background-color: #ffffff;
                border: 1px solid #d0d0d0;
                border-radius: 4px;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #c0c0c0;
                border-radius: 5px;
                margin-top: 0.5em;
                padding-top: 0.5em;
                background-color: #fafafa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #2c3e50;
            }
            QPushButton {
                background-color: #3498db;
                border: none;
                color: white;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover { background-color: #2980b9; }
            QPushButton:pressed { background-color: #1c6ea4; }
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9f9f9;
                gridline-color: #e0e0e0;
                selection-background-color: #b0d4f0;
                border: 1px solid #d0d0d0;
                border-radius: 3px;
            }
            QTableWidget::item { padding: 3px; }
            QHeaderView::section {
                background-color: #e8e8e8;
                padding: 4px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
                font-size: 11px;
            }
            QListWidget {
                border: 1px solid #d0d0d0;
                border-radius: 3px;
                background: white;
                font-size: 12px;
            }
            QTextEdit {
                border: 1px solid #d0d0d0;
                border-radius: 3px;
                background: white;
                font-family: Consolas, monospace;
                font-size: 11px;
            }
            QLineEdit {
                border: 1px solid #d0d0d0;
                border-radius: 3px;
                padding: 4px;
                background: white;
            }
            QLineEdit:focus { border: 1px solid #3498db; }
            QTabWidget::pane {
                border: 1px solid #d0d0d0;
                border-radius: 3px;
                background: white;
            }
            QTabBar::tab {
                background: #e8e8e8;
                border: 1px solid #d0d0d0;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                padding: 5px 12px;
                margin-right: 2px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom: 2px solid #3498db;
            }
            QTabBar::tab:hover { background: #d0d0d0; }
            QStatusBar {
                background-color: #f0f0f0;
                color: #2c3e50;
                font-size: 11px;
            }
            QLabel {
                color: #2c3e50;
                font-weight: bold;
            }
        """
        self.setStyleSheet(style)

    # ---------- 文件操作 ----------
    def add_files(self):
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "选择诊断日志文件", "",
            "Text Files (*.txt);;Excel Files (*.xlsx);;All Files (*)"
        )
        if file_paths:
            for path in file_paths:
                self.load_file(path)

    def load_file(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        content = None
        try:
            if ext == '.xlsx':
                wb = openpyxl.load_workbook(file_path, data_only=True)
                target_sheet = None
                if "Diagnoseprotokoll" in wb.sheetnames:
                    target_sheet = "Diagnoseprotokoll"
                else:
                    for name in wb.sheetnames:
                        if "protokoll" in name.lower():
                            target_sheet = name
                            break
                if target_sheet is None:
                    target_sheet = wb.sheetnames[0]
                    self.statusBar().showMessage(f"未找到 'Diagnoseprotokoll'，使用第一个工作表: {target_sheet}")
                ws = wb[target_sheet]
                lines = []
                for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                    cell_val = row[0]
                    if cell_val is not None:
                        lines.append(str(cell_val))
                if len(lines) == 1 and '\n' in lines[0]:
                    content = lines[0]
                else:
                    content = "\n".join(lines)
                self.statusBar().showMessage(f"从Excel读取到 {len(lines)} 行数据")
            else:
                encodings = ['utf-8-sig', 'utf-8', 'gbk', 'latin-1']
                for enc in encodings:
                    try:
                        with open(file_path, 'r', encoding=enc) as f:
                            content = f.read()
                        break
                    except UnicodeDecodeError:
                        continue
                if content is None:
                    QMessageBox.critical(self, "错误", f"无法解码文件 {file_path}，请检查编码。")
                    return
                content = content.lstrip('\ufeff')
        except Exception as e:
            QMessageBox.critical(self, "读取错误", f"读取文件失败: {str(e)}")
            return

        if content is None:
            return

        try:
            data = parse_dtc_text(content)
            if not data['main'] and not data['perf']:
                preview = content[:300] + "..." if len(content) > 300 else content
                QMessageBox.warning(self, "解析警告",
                    f"文件 {file_path} 未解析到任何DTC数据。\n"
                    f"内容预览（前300字符）：\n{preview}\n"
                    f"请确认内容包含 'IPN_MAIN'/'IPN_PERF' 和阶段标题（如 Preprocess）。")
                return
            name = os.path.basename(file_path)
            item = FileItem(name, data)
            self.files.append(item)
            self.file_list.addItem(name)
            self.update_tables()
            self.update_single_tabs()
            self.statusBar().showMessage(f"已加载文件: {name}")
        except Exception as e:
            QMessageBox.critical(self, "解析错误", f"解析文件失败: {str(e)}\n\n{traceback.format_exc()}")

    def remove_selected(self):
        selected = self.file_list.selectedItems()
        if not selected:
            return
        rows = sorted([self.file_list.row(item) for item in selected], reverse=True)
        for row in rows:
            self.file_list.takeItem(row)
            del self.files[row]
        self.update_tables()
        self.update_single_tabs()
        self.statusBar().showMessage(f"已移除 {len(rows)} 个文件")

    def clear_all(self):
        self.file_list.clear()
        self.files.clear()
        self.paste_edit.clear()
        self.update_tables()
        self.update_single_tabs()
        self.statusBar().showMessage("已清空所有文件")

    def parse_paste(self):
        text = self.paste_edit.toPlainText().strip()
        if not text:
            QMessageBox.information(self, "提示", "请先粘贴内容")
            return
        blocks = [b.strip() for b in text.split('---') if b.strip()]
        if not blocks:
            QMessageBox.warning(self, "解析错误", "未找到有效内容块，请用 '---' 分隔多个文件")
            return
        loaded = 0
        for idx, block in enumerate(blocks):
            try:
                data = parse_dtc_text(block)
                if not data['main'] and not data['perf']:
                    continue
                name = f"粘贴块 {idx+1}"
                item = FileItem(name, data)
                self.files.append(item)
                self.file_list.addItem(name)
                loaded += 1
            except Exception as e:
                QMessageBox.warning(self, "解析警告", f"解析第{idx+1}块失败: {str(e)}")
        if loaded > 0:
            self.update_tables()
            self.update_single_tabs()
            self.statusBar().showMessage(f"成功解析 {loaded} 个粘贴块")
        else:
            QMessageBox.warning(self, "解析失败", "未能解析任何有效数据，请检查格式。")

    # ---------- 表格更新 ----------
    def update_tables(self):
        self.update_overview_table(self.overview_main_table, 'main')
        self.update_overview_table(self.overview_perf_table, 'perf')
        QTimer.singleShot(50, self.unify_overview_widths)

    def update_overview_table(self, table_widget, module):
        table_widget.clear()
        if not self.files:
            table_widget.setRowCount(0)
            table_widget.setColumnCount(0)
            return

        num_files = len(self.files)
        total_cols = 3 + 3 * num_files
        table_widget.setColumnCount(total_cols)
        
        # 使用数据区域模拟双行表头：第0行文件名（合并三列），第1行阶段名，之后是数据
        # 先收集所有DTC
        order = []
        for f in self.files:
            for dtc_id, _, _, _, _, _ in f.data[module]:
                if dtc_id not in order:
                    order.append(dtc_id)
        data_row_count = len(order)
        table_widget.setRowCount(2 + data_row_count)
        
        # 设置第0行：文件名合并
        fixed_headers = ['DTC编号', '短名称', '描述']
        for col in range(3):
            item = QTableWidgetItem(fixed_headers[col])
            item.setBackground(QBrush(QColor(200, 200, 200)))  # 浅灰背景
            item.setTextAlignment(Qt.AlignCenter)
            table_widget.setItem(0, col, item)
        
        # 从第3列开始，每个文件占3列，合并单元格显示文件名
        col = 3
        for f_idx, f in enumerate(self.files):
            item = QTableWidgetItem(f.name)
            item.setBackground(QBrush(self.get_file_color(f_idx)))
            item.setTextAlignment(Qt.AlignCenter)
            table_widget.setItem(0, col, item)
            if 3 * num_files > 1:
                table_widget.setSpan(0, col, 1, 3)
            col += 3
        
        # 设置第1行：阶段名（不合并）
        stage_names = ['Preprocess', 'Before', 'Postprocess']
        for col in range(3):
            item = QTableWidgetItem("")
            item.setBackground(QBrush(QColor(220, 220, 220)))
            table_widget.setItem(1, col, item)
        col = 3
        for f_idx, f in enumerate(self.files):
            for stage in stage_names:
                item = QTableWidgetItem(stage)
                item.setBackground(QBrush(self.get_file_color(f_idx)))
                item.setTextAlignment(Qt.AlignCenter)
                table_widget.setItem(1, col, item)
                col += 1
        
        # 填充数据行（从第2行开始）
        for row, dtc_id in enumerate(order):
            real_row = row + 2
            # 前三列
            table_widget.setItem(real_row, 0, QTableWidgetItem(dtc_id))
            short = desc = ''
            for f in self.files:
                for d in f.data[module]:
                    if d[0] == dtc_id:
                        short, desc = d[1], d[2]
                        break
                if short:
                    break
            table_widget.setItem(real_row, 1, QTableWidgetItem(short))
            table_widget.setItem(real_row, 2, QTableWidgetItem(desc))
            
            col = 3
            for f in self.files:
                pre = before = post = ''
                for d in f.data[module]:
                    if d[0] == dtc_id:
                        pre, before, post = d[3], d[4], d[5]
                        break
                for status, off in [(pre, 0), (before, 1), (post, 2)]:
                    item = QTableWidgetItem(status)
                    color = self.get_color(status)
                    if color:
                        item.setBackground(QBrush(color))
                    table_widget.setItem(real_row, col + off, item)
                col += 3
        
        # 设置行高
        table_widget.setRowHeight(0, 30)
        table_widget.setRowHeight(1, 25)

    def get_file_color(self, index):
        colors = [
            QColor(200, 230, 255),  # 淡蓝
            QColor(200, 255, 200),  # 淡绿
            QColor(255, 230, 200),  # 淡橙
            QColor(255, 200, 230),  # 淡粉
            QColor(230, 200, 255),  # 淡紫
            QColor(200, 255, 255),  # 淡青
        ]
        return colors[index % len(colors)]

    def unify_overview_widths(self):
        tables = [self.overview_main_table, self.overview_perf_table]
        if tables[0].columnCount() == 0:
            return
        # 先自适应
        for table in tables:
            table.resizeColumnsToContents()
        col_count = tables[0].columnCount()
        max_widths = [0] * col_count
        for table in tables:
            for col in range(col_count):
                width = table.columnWidth(col)
                if width > max_widths[col]:
                    max_widths[col] = width
        for table in tables:
            for col in range(col_count):
                table.setColumnWidth(col, max_widths[col])

    def get_color(self, status):
        if status == '0x2f':
            return QColor(255, 200, 200)
        elif status == '0x2e':
            return QColor(200, 255, 200)
        return None

    def update_single_tabs(self):
        while self.tab_widget.count() > 1:
            self.tab_widget.removeTab(1)

        for f in self.files:
            tab = QWidget()
            layout = QVBoxLayout(tab)
            main_table = QTableWidget()
            perf_table = QTableWidget()
            main_table.setSortingEnabled(False)
            perf_table.setSortingEnabled(False)
            font = QFont("Segoe UI", 9)
            main_table.setFont(font)
            perf_table.setFont(font)
            main_table.setAlternatingRowColors(True)
            perf_table.setAlternatingRowColors(True)
            main_table.verticalHeader().setVisible(False)
            perf_table.verticalHeader().setVisible(False)
            layout.addWidget(QLabel("IPN_MAIN"))
            layout.addWidget(main_table)
            layout.addWidget(QLabel("IPN_PERF"))
            layout.addWidget(perf_table)
            self.populate_single_table(main_table, f.data['main'])
            self.populate_single_table(perf_table, f.data['perf'])
            self.unify_single_table_widths(main_table, perf_table)
            self.tab_widget.addTab(tab, f.name)

    def populate_single_table(self, table, data_list):
        if not data_list:
            table.setRowCount(0)
            table.setColumnCount(0)
            return
        headers = ['DTC编号', '短名称', '描述', 'Preprocess', 'Before Stimulation', 'Postprocess']
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(data_list))
        for row, (dtc, short, desc, pre, before, post) in enumerate(data_list):
            table.setItem(row, 0, QTableWidgetItem(dtc))
            table.setItem(row, 1, QTableWidgetItem(short))
            table.setItem(row, 2, QTableWidgetItem(desc))
            for off, status in enumerate([pre, before, post], start=3):
                item = QTableWidgetItem(status)
                color = self.get_color(status)
                if color:
                    item.setBackground(QBrush(color))
                table.setItem(row, off, item)
        table.resizeColumnsToContents()

    def unify_single_table_widths(self, table1, table2):
        if table1.columnCount() == 0 or table2.columnCount() == 0:
            return
        if table1.columnCount() != table2.columnCount():
            return
        table1.resizeColumnsToContents()
        table2.resizeColumnsToContents()
        col_count = table1.columnCount()
        max_widths = [0] * col_count
        for table in [table1, table2]:
            for col in range(col_count):
                width = table.columnWidth(col)
                if width > max_widths[col]:
                    max_widths[col] = width
        for table in [table1, table2]:
            for col in range(col_count):
                table.setColumnWidth(col, max_widths[col])

    # ---------- 搜索 ----------
    def on_search(self, text):
        self.filter_table(self.overview_main_table, text)
        self.filter_table(self.overview_perf_table, text)

    def filter_table(self, table_widget, text):
        if not text:
            for row in range(table_widget.rowCount()):
                table_widget.setRowHidden(row, False)
            return
        text = text.lower()
        for row in range(table_widget.rowCount()):
            if row < 2:  # 跳过表头行
                continue
            match = False
            dtc_item = table_widget.item(row, 0)
            desc_item = table_widget.item(row, 2)
            if dtc_item and text in dtc_item.text().lower():
                match = True
            elif desc_item and text in desc_item.text().lower():
                match = True
            table_widget.setRowHidden(row, not match)

    # ---------- 导出 ----------
    def export_current(self):
        idx = self.tab_widget.currentIndex()
        if idx == 0:
            self.export_overview()
        else:
            tab = self.tab_widget.widget(idx)
            self.export_single_file(tab, self.tab_widget.tabText(idx))

    def export_overview(self):
        main_data = self.get_visible_table_data(self.overview_main_table)
        perf_data = self.get_visible_table_data(self.overview_perf_table)
        if not main_data and not perf_data:
            QMessageBox.information(self, "提示", "没有可见数据可导出")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "保存Excel文件", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            if main_data:
                ws = wb.create_sheet("IPN_MAIN")
                self.write_table_to_sheet(ws, main_data)
            if perf_data:
                ws = wb.create_sheet("IPN_PERF")
                self.write_table_to_sheet(ws, perf_data)
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"已保存到 {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def export_single_file(self, tab, tab_name):
        tables = tab.findChildren(QTableWidget)
        if len(tables) < 2:
            QMessageBox.information(self, "提示", "该标签页没有数据")
            return
        main_table, perf_table = tables[0], tables[1]
        main_data = self.get_visible_table_data(main_table)
        perf_data = self.get_visible_table_data(perf_table)
        if not main_data and not perf_data:
            QMessageBox.information(self, "提示", "没有可见数据可导出")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "保存Excel文件", f"{tab_name}.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            if main_data:
                ws = wb.create_sheet("IPN_MAIN")
                self.write_table_to_sheet(ws, main_data)
            if perf_data:
                ws = wb.create_sheet("IPN_PERF")
                self.write_table_to_sheet(ws, perf_data)
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"已保存到 {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def get_visible_table_data(self, table_widget):
        if table_widget.rowCount() == 0:
            return None
        col_count = table_widget.columnCount()
        # 构建表头：从第0行和第1行组合
        headers = []
        for col in range(3):
            item0 = table_widget.item(0, col)
            headers.append(item0.text() if item0 else f"Col{col}")
        col = 3
        while col < col_count:
            item0 = table_widget.item(0, col)
            file_name = item0.text() if item0 else "File"
            for off in range(3):
                item1 = table_widget.item(1, col + off)
                stage = item1.text() if item1 else "Stage"
                headers.append(f"{file_name}\n{stage}")
            col += 3
        data = []
        for row in range(2, table_widget.rowCount()):
            if table_widget.isRowHidden(row):
                continue
            row_data = []
            for col in range(col_count):
                item = table_widget.item(row, col)
                if item:
                    text = item.text()
                    bg = item.background().color()
                    color_hex = bg.name() if bg.isValid() else None
                    row_data.append((text, color_hex))
                else:
                    row_data.append(('', None))
            data.append(row_data)
        return {'headers': headers, 'data': data}

    def write_table_to_sheet(self, ws, table_data):
        headers = table_data['headers']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        for row_idx, row in enumerate(table_data['data'], start=2):
            for col_idx, (text, color_hex) in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=text)
                if color_hex:
                    fill = PatternFill(start_color=color_hex[1:], end_color=color_hex[1:], fill_type="solid")
                    cell.fill = fill

    # ---------- 拖拽 ----------
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path:
                self.load_file(file_path)
        event.acceptProposedAction()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = DtcCompareApp()
    window.show()
    sys.exit(app.exec_())