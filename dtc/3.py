import sys
import re
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QFileDialog, QTabWidget, QLineEdit, QLabel,
    QListWidget, QAbstractItemView, QMessageBox, QTextEdit,
    QGroupBox, QFrame
)
from PyQt5.QtCore import Qt, QUrl
from PyQt5.QtGui import QColor, QBrush, QDragEnterEvent, QDropEvent
import openpyxl
from openpyxl.styles import PatternFill

# ---------- 改进后的解析函数 ----------
def parse_dtc_text(text):
    """
    解析与dtc.txt格式相同的文本。
    返回：{'main': [(id, short, desc, pre, before, post), ...], 'perf': [...]}
    """
    lines = text.splitlines()
    phases = ['Preprocess', 'Before Stimulation', 'Postprocess']
    phase_data = {phase: {'main': [], 'perf': []} for phase in phases}
    
    current_phase = None
    current_module = None
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        # 检测阶段标题 (如 "(514.81480) Preprocess")
        m = re.match(r'^\(\d+\.\d+\)\s+(.+)$', line)
        if m:
            current_phase = m.group(1).strip()
            i += 1
            continue
        
        # 检测模块
        if line.startswith('IPN_MAIN'):
            current_module = 'main'
            i += 1
            continue
        elif line.startswith('IPN_PERF'):
            current_module = 'perf'
            i += 1
            continue
        
        # 检测DTC条目 (0x开头的6位十六进制)
        if re.match(r'^0x[0-9A-Fa-f]{6}$', line):
            dtc_id = line
            short_name = ''
            desc = ''
            status = ''
            j = i + 1
            # 在DTC块中读取键值对
            while j < len(lines):
                line_j = lines[j].strip()
                if not line_j:
                    j += 1
                    continue
                # 遇到下一个DTC或新模块/阶段则停止
                if re.match(r'^0x[0-9A-Fa-f]{6}$', line_j):
                    break
                if line_j.startswith('IPN_MAIN') or line_j.startswith('IPN_PERF') or re.match(r'^\(\d+\.\d+\)', line_j):
                    break
                
                # 处理键：ShortName, Description, Status
                if line_j.startswith('ShortName'):
                    # 值通常在下一行
                    if j + 1 < len(lines):
                        val = lines[j+1].strip()
                        # 如果值可能是空行，跳过
                        if val:
                            short_name = val
                    j += 2  # 跳过键行和值行
                    continue
                elif line_j.startswith('Description'):
                    if j + 1 < len(lines):
                        val = lines[j+1].strip()
                        if val:
                            desc = val
                    j += 2
                    continue
                elif line_j.startswith('Status'):
                    if j + 1 < len(lines):
                        val = lines[j+1].strip()
                        if val:
                            status = val
                    j += 2
                    continue
                else:
                    # 其他行（如空缩进或无关内容）跳过
                    j += 1
                    continue
            
            # 记录DTC数据
            if current_phase and current_module and current_phase in phase_data:
                phase_data[current_phase][current_module].append((dtc_id, short_name, desc, status))
            i = j
            continue
        i += 1

    # 构建DTC顺序（按Preprocess中的出现顺序，若没有则按Before，再按Post）
    order_main = []
    order_perf = []
    for phase in phases:
        for module in ['main', 'perf']:
            for dtc_id, _, _, _ in phase_data[phase][module]:
                if module == 'main' and dtc_id not in order_main:
                    order_main.append(dtc_id)
                elif module == 'perf' and dtc_id not in order_perf:
                    order_perf.append(dtc_id)
    
    # 整合三个阶段的状态
    file_data = {'main': {}, 'perf': {}}
    for phase in phases:
        for module in ['main', 'perf']:
            for dtc_id, short, desc, status in phase_data[phase][module]:
                if dtc_id not in file_data[module]:
                    file_data[module][dtc_id] = {'short': short, 'desc': desc, 'pre': '', 'before': '', 'post': ''}
                if phase == 'Preprocess':
                    file_data[module][dtc_id]['pre'] = status
                elif phase == 'Before Stimulation':
                    file_data[module][dtc_id]['before'] = status
                elif phase == 'Postprocess':
                    file_data[module][dtc_id]['post'] = status
    
    # 生成最终列表
    main_list = []
    for dtc in order_main:
        info = file_data['main'][dtc]
        main_list.append((dtc, info['short'], info['desc'], info['pre'], info['before'], info['post']))
    
    perf_list = []
    for dtc in order_perf:
        info = file_data['perf'][dtc]
        perf_list.append((dtc, info['short'], info['desc'], info['pre'], info['before'], info['post']))
    
    return {'main': main_list, 'perf': perf_list}


class FileItem:
    def __init__(self, name, data):
        self.name = name
        self.data = data


class DtcCompareApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DTC 对比分析工具")
        self.setGeometry(100, 100, 1400, 800)
        self.files = []
        self.initUI()
        self.apply_styles()

    def initUI(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(5)
        main_layout.setContentsMargins(5, 5, 5, 5)

        # ---------- 顶部控制区域（紧凑） ----------
        top_frame = QFrame()
        top_frame.setFrameShape(QFrame.StyledPanel)
        top_layout = QHBoxLayout(top_frame)
        top_layout.setSpacing(5)

        # 文件管理组
        file_group = QGroupBox("文件")
        file_group.setMaximumHeight(100)
        file_layout = QHBoxLayout(file_group)
        self.btn_add = QPushButton("添加")
        self.btn_add.clicked.connect(self.add_files)
        self.btn_remove = QPushButton("移除")
        self.btn_remove.clicked.connect(self.remove_selected)
        self.btn_clear = QPushButton("清空")
        self.btn_clear.clicked.connect(self.clear_all)
        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.file_list.setMaximumHeight(80)
        file_layout.addWidget(self.btn_add)
        file_layout.addWidget(self.btn_remove)
        file_layout.addWidget(self.btn_clear)
        file_layout.addWidget(self.file_list, 1)
        top_layout.addWidget(file_group, 1)

        # 粘贴区域（折叠式）
        paste_group = QGroupBox("▼ 粘贴解析")
        paste_group.setCheckable(True)
        paste_group.setChecked(False)
        paste_group.setMaximumHeight(30)
        paste_group.toggled.connect(lambda checked: paste_group.setMaximumHeight(120 if checked else 30))
        paste_layout = QVBoxLayout(paste_group)
        self.paste_edit = QTextEdit()
        self.paste_edit.setPlaceholderText("在此粘贴dtc.txt格式内容... 多个文件用 '---' 分隔")
        self.paste_edit.setMaximumHeight(60)
        self.btn_paste_parse = QPushButton("解析粘贴")
        self.btn_paste_parse.clicked.connect(self.parse_paste)
        paste_layout.addWidget(self.paste_edit)
        paste_layout.addWidget(self.btn_paste_parse)
        top_layout.addWidget(paste_group, 1)

        main_layout.addWidget(top_frame)

        # ---------- 搜索栏 ----------
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索:"))
        self.search_line = QLineEdit()
        self.search_line.setPlaceholderText("输入DTC编号或描述关键字...")
        self.search_line.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_line)
        main_layout.addLayout(search_layout)

        # ---------- 标签页 ----------
        self.tab_widget = QTabWidget()
        self.tab_widget.setTabsClosable(False)
        self.overview_tab = QWidget()
        self.setup_overview_tab()
        self.tab_widget.addTab(self.overview_tab, "总览对比")
        main_layout.addWidget(self.tab_widget, 1)

        # ---------- 底部导出 ----------
        export_layout = QHBoxLayout()
        self.btn_export = QPushButton("导出当前页面为 Excel")
        self.btn_export.clicked.connect(self.export_current)
        export_layout.addStretch()
        export_layout.addWidget(self.btn_export)
        main_layout.addLayout(export_layout)

        self.statusBar().showMessage("就绪")
        self.setAcceptDrops(True)

    def setup_overview_tab(self):
        layout = QVBoxLayout(self.overview_tab)
        layout.setSpacing(2)
        self.overview_main_table = QTableWidget()
        self.overview_perf_table = QTableWidget()
        self.overview_main_table.setSortingEnabled(False)
        self.overview_perf_table.setSortingEnabled(False)
        layout.addWidget(QLabel("IPN_MAIN"))
        layout.addWidget(self.overview_main_table)
        layout.addWidget(QLabel("IPN_PERF"))
        layout.addWidget(self.overview_perf_table)

    def apply_styles(self):
        style = """
            QMainWindow { background-color: #f0f0f0; }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ccc;
                border-radius: 5px;
                margin-top: 0.5em;
                padding-top: 0.5em;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QPushButton {
                background-color: #4CAF50;
                border: none;
                color: white;
                padding: 5px 10px;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9f9f9;
                gridline-color: #ddd;
                selection-background-color: #b0d4f0;
            }
            QTableWidget::item { padding: 2px; }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 4px;
                border: 1px solid #ccc;
                font-weight: bold;
            }
            QListWidget {
                border: 1px solid #ccc;
                border-radius: 3px;
                background: white;
            }
            QTextEdit {
                border: 1px solid #ccc;
                border-radius: 3px;
                background: white;
            }
            QLineEdit {
                border: 1px solid #ccc;
                border-radius: 3px;
                padding: 3px;
                background: white;
            }
        """
        self.setStyleSheet(style)

    # ---------- 文件操作 ----------
    def add_files(self):
        file_paths, _ = QFileDialog.getOpenFileNames(self, "选择诊断日志文件", "", "Text Files (*.txt);;All Files (*)")
        if file_paths:
            for path in file_paths:
                self.load_file(path)

    def load_file(self, file_path):
        encodings = ['utf-8', 'gbk', 'latin-1']
        content = None
        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        if content is None:
            QMessageBox.critical(self, "错误", f"无法解码文件 {file_path}，请检查编码。")
            return
        try:
            data = parse_dtc_text(content)
            if not data['main'] and not data['perf']:
                QMessageBox.warning(self, "解析警告", f"文件 {file_path} 未解析到任何DTC数据。")
                return
            name = os.path.basename(file_path)
            item = FileItem(name, data)
            self.files.append(item)
            self.file_list.addItem(name)
            self.update_tables()
            self.update_single_tabs()
            self.statusBar().showMessage(f"已加载文件: {name}")
        except Exception as e:
            QMessageBox.critical(self, "解析错误", f"解析文件失败: {str(e)}")

    def remove_selected(self):
        selected = self.file_list.selectedItems()
        if not selected:
            return
        rows = sorted([self.file_list.row(item) for item in selected], reverse=True)
        for row in rows:
            self.file_list.takeItem(row)
            del self.files[row]
        self.update_tables()
        self.update_single_tabs()
        self.statusBar().showMessage(f"已移除 {len(rows)} 个文件")

    def clear_all(self):
        self.file_list.clear()
        self.files.clear()
        self.paste_edit.clear()
        self.update_tables()
        self.update_single_tabs()
        self.statusBar().showMessage("已清空所有文件")

    def parse_paste(self):
        text = self.paste_edit.toPlainText().strip()
        if not text:
            QMessageBox.information(self, "提示", "请先粘贴内容")
            return
        blocks = [b.strip() for b in text.split('---') if b.strip()]
        if not blocks:
            QMessageBox.warning(self, "解析错误", "未找到有效内容块，请用 '---' 分隔多个文件")
            return
        loaded = 0
        for idx, block in enumerate(blocks):
            try:
                data = parse_dtc_text(block)
                if not data['main'] and not data['perf']:
                    continue
                name = f"粘贴块 {idx+1}"
                item = FileItem(name, data)
                self.files.append(item)
                self.file_list.addItem(name)
                loaded += 1
            except Exception as e:
                QMessageBox.warning(self, "解析警告", f"解析第{idx+1}块失败: {str(e)}")
        if loaded > 0:
            self.update_tables()
            self.update_single_tabs()
            self.statusBar().showMessage(f"成功解析 {loaded} 个粘贴块")
        else:
            QMessageBox.warning(self, "解析失败", "未能解析任何有效数据，请检查格式。")

    # ---------- 表格更新 ----------
    def update_tables(self):
        self.update_overview_table(self.overview_main_table, 'main')
        self.update_overview_table(self.overview_perf_table, 'perf')

    def update_overview_table(self, table_widget, module):
        table_widget.clear()
        if not self.files:
            table_widget.setRowCount(0)
            table_widget.setColumnCount(0)
            return

        num_files = len(self.files)
        col_count = 3 + 3 * num_files
        headers = ['DTC编号', '短名称', '描述']
        for f in self.files:
            headers.append(f"{f.name}-Pre")
            headers.append(f"{f.name}-Before")
            headers.append(f"{f.name}-Post")
        table_widget.setColumnCount(col_count)
        table_widget.setHorizontalHeaderLabels(headers)

        # 收集DTC（按第一个文件顺序）
        order = []
        for f in self.files:
            for dtc_id, _, _, _, _, _ in f.data[module]:
                if dtc_id not in order:
                    order.append(dtc_id)

        table_widget.setRowCount(len(order))
        for row, dtc_id in enumerate(order):
            # DTC, short, desc
            table_widget.setItem(row, 0, QTableWidgetItem(dtc_id))
            short = desc = ''
            for f in self.files:
                for d in f.data[module]:
                    if d[0] == dtc_id:
                        short, desc = d[1], d[2]
                        break
                if short:
                    break
            table_widget.setItem(row, 1, QTableWidgetItem(short))
            table_widget.setItem(row, 2, QTableWidgetItem(desc))

            col = 3
            for f in self.files:
                pre = before = post = ''
                for d in f.data[module]:
                    if d[0] == dtc_id:
                        pre, before, post = d[3], d[4], d[5]
                        break
                for status, off in [(pre, 0), (before, 1), (post, 2)]:
                    item = QTableWidgetItem(status)
                    color = self.get_color(status)
                    if color:
                        item.setBackground(QBrush(color))
                    table_widget.setItem(row, col + off, item)
                col += 3

        table_widget.resizeColumnsToContents()

    def get_color(self, status):
        if status == '0x2f':
            return QColor(255, 200, 200)
        elif status == '0x2e':
            return QColor(200, 255, 200)
        return None

    def update_single_tabs(self):
        while self.tab_widget.count() > 1:
            self.tab_widget.removeTab(1)

        for f in self.files:
            tab = QWidget()
            layout = QVBoxLayout(tab)
            main_table = QTableWidget()
            perf_table = QTableWidget()
            main_table.setSortingEnabled(False)
            perf_table.setSortingEnabled(False)
            layout.addWidget(QLabel("IPN_MAIN"))
            layout.addWidget(main_table)
            layout.addWidget(QLabel("IPN_PERF"))
            layout.addWidget(perf_table)
            self.populate_single_table(main_table, f.data['main'])
            self.populate_single_table(perf_table, f.data['perf'])
            self.tab_widget.addTab(tab, f.name)

    def populate_single_table(self, table, data_list):
        if not data_list:
            table.setRowCount(0)
            table.setColumnCount(0)
            return
        headers = ['DTC编号', '短名称', '描述', 'Preprocess', 'Before', 'Postprocess']
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(data_list))
        for row, (dtc, short, desc, pre, before, post) in enumerate(data_list):
            table.setItem(row, 0, QTableWidgetItem(dtc))
            table.setItem(row, 1, QTableWidgetItem(short))
            table.setItem(row, 2, QTableWidgetItem(desc))
            for off, status in enumerate([pre, before, post], start=3):
                item = QTableWidgetItem(status)
                color = self.get_color(status)
                if color:
                    item.setBackground(QBrush(color))
                table.setItem(row, off, item)
        table.resizeColumnsToContents()

    # ---------- 搜索 ----------
    def on_search(self, text):
        self.filter_table(self.overview_main_table, text)
        self.filter_table(self.overview_perf_table, text)

    def filter_table(self, table, text):
        if not text:
            for row in range(table.rowCount()):
                table.setRowHidden(row, False)
            return
        text = text.lower()
        for row in range(table.rowCount()):
            match = False
            dtc_item = table.item(row, 0)
            desc_item = table.item(row, 2)
            if dtc_item and text in dtc_item.text().lower():
                match = True
            elif desc_item and text in desc_item.text().lower():
                match = True
            table.setRowHidden(row, not match)

    # ---------- 导出 ----------
    def export_current(self):
        idx = self.tab_widget.currentIndex()
        if idx == 0:
            self.export_overview()
        else:
            tab = self.tab_widget.widget(idx)
            self.export_single_file(tab, self.tab_widget.tabText(idx))

    def export_overview(self):
        main_data = self.get_visible_table_data(self.overview_main_table)
        perf_data = self.get_visible_table_data(self.overview_perf_table)
        if not main_data and not perf_data:
            QMessageBox.information(self, "提示", "没有可见数据可导出")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "保存Excel文件", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            if main_data:
                ws = wb.create_sheet("IPN_MAIN")
                self.write_table_to_sheet(ws, main_data)
            if perf_data:
                ws = wb.create_sheet("IPN_PERF")
                self.write_table_to_sheet(ws, perf_data)
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"已保存到 {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def export_single_file(self, tab, tab_name):
        tables = tab.findChildren(QTableWidget)
        if len(tables) < 2:
            QMessageBox.information(self, "提示", "该标签页没有数据")
            return
        main_table, perf_table = tables[0], tables[1]
        main_data = self.get_visible_table_data(main_table)
        perf_data = self.get_visible_table_data(perf_table)
        if not main_data and not perf_data:
            QMessageBox.information(self, "提示", "没有可见数据可导出")
            return
        file_path, _ = QFileDialog.getSaveFileName(self, "保存Excel文件", f"{tab_name}.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            if main_data:
                ws = wb.create_sheet("IPN_MAIN")
                self.write_table_to_sheet(ws, main_data)
            if perf_data:
                ws = wb.create_sheet("IPN_PERF")
                self.write_table_to_sheet(ws, perf_data)
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"已保存到 {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def get_visible_table_data(self, table):
        if table.rowCount() == 0:
            return None
        headers = [table.horizontalHeaderItem(c).text() if table.horizontalHeaderItem(c) else f"Col{c}" for c in range(table.columnCount())]
        data = []
        for row in range(table.rowCount()):
            if table.isRowHidden(row):
                continue
            row_data = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    text = item.text()
                    bg = item.background().color()
                    color_hex = bg.name() if bg.isValid() else None
                    row_data.append((text, color_hex))
                else:
                    row_data.append(('', None))
            data.append(row_data)
        return {'headers': headers, 'data': data}

    def write_table_to_sheet(self, ws, table_data):
        headers = table_data['headers']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        for row_idx, row in enumerate(table_data['data'], start=2):
            for col_idx, (text, color_hex) in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=text)
                if color_hex:
                    fill = PatternFill(start_color=color_hex[1:], end_color=color_hex[1:], fill_type="solid")
                    cell.fill = fill

    # ---------- 拖拽 ----------
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path:
                self.load_file(file_path)
        event.acceptProposedAction()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = DtcCompareApp()
    window.show()
    sys.exit(app.exec_())